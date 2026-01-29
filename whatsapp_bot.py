import pywhatkit as kit
import openpyxl
import requests
import re
import time


def buscar_endereco_por_cep(cep):
    cep = re.sub(r"\D", "", str(cep))
    if len(cep) != 8:
        return None

    try:
        r = requests.get(f"https://viacep.com.br/ws/{cep}/json/", timeout=10)
        dados = r.json()
        if dados.get("erro"):
            return None

        return {
            "bairro": dados.get("bairro") or dados.get("localidade", ""),
            "cidade": dados.get("localidade", ""),
            "uf": dados.get("uf", "")
        }
    except:
        return None

def enviar_mensagem(numero, mensagem):
    kit.sendwhatmsg_instantly(
        phone_no=f"+55{numero}",
        message=mensagem,
        wait_time=12,
        tab_close=True,
        close_time=3
    )

def enviar_mensagens(caminho_planilha):
    wb = openpyxl.load_workbook(caminho_planilha)
    ws = wb.active

  
    cab = {cell.value.strip().lower(): i for i, cell in enumerate(ws[1])}

   
    def garantir_coluna(nome):
        if nome not in cab:
            ws.cell(1, ws.max_column + 1, nome)
            cab[nome] = ws.max_column - 1

    garantir_coluna("telefone2")
    garantir_coluna("complemento")
    garantir_coluna("status_envio")
    garantir_coluna("telefone_usado")

    for idx, linha in enumerate(ws.iter_rows(min_row=2), start=2):
        nome = linha[cab["cliente"]].value
        tel1 = linha[cab["telefone"]].value
        tel2 = linha[cab["telefone2"]].value
        servico = linha[cab["serviço"]].value
        produto = linha[cab["produto"]].value
        logradouro = linha[cab["logradouro"]].value
        numero = linha[cab["número"]].value
        complemento = linha[cab["complemento"]].value
        cep = linha[cab["cep"]].value

        if not nome or not cep:
            ws.cell(idx, cab["status_envio"] + 1, "Dados incompletos")
            ws.cell(idx, cab["telefone_usado"] + 1, "-")
            continue

        endereco = buscar_endereco_por_cep(cep)
        if not endereco:
            ws.cell(idx, cab["status_envio"] + 1, "Erro endereço")
            ws.cell(idx, cab["telefone_usado"] + 1, "-")
            continue

        bairro = endereco["bairro"]
        cidade = endereco["cidade"]
        uf = endereco["uf"]

        
        endereco_linha = f"{logradouro}, nº {numero}"
        if complemento:
            endereco_linha += f" - {complemento}"

        # 
        mensagem = (
            f"Olá {nome}, tudo bem?\n\n"
            f"Estamos entrando em contato referente ao serviço de {servico}.\n\n"
            f"Produto: {produto}\n\n"
            f"Endereço:\n"
            f"{endereco_linha}\n"
            f"{bairro} - {cidade}/{uf}\n"
            f"CEP: {cep}\n\n"
            f"Poderia confirmar, por gentileza, se as informações estão corretas?\n\n"
            f"Atenciosamente."
        )

        tentou = False

        # 🔹 TELEFONE 1
        if tel1:
            tel1 = re.sub(r"\D", "", str(tel1))
            print(f"📞 Tentando telefone 1 ({nome})")
            enviar_mensagem(tel1, mensagem)
            ws.cell(idx, cab["telefone_usado"] + 1, "Telefone 1")
            tentou = True
            time.sleep(15)

        # 🔹 TELEFONE 2
        if tel2:
            tel2 = re.sub(r"\D", "", str(tel2))
            print(f"📞 Tentando telefone 2 ({nome})")
            enviar_mensagem(tel2, mensagem)
            ws.cell(idx, cab["telefone_usado"] + 1, "Telefone 2")
            tentou = True
            time.sleep(15)

        if tentou:
            ws.cell(idx, cab["status_envio"] + 1, "Enviado")
        else:
            ws.cell(idx, cab["status_envio"] + 1, "Sem telefone")
            ws.cell(idx, cab["telefone_usado"] + 1, "-")

    wb.save(caminho_planilha)
    print("✅ Planilha atualizada corretamente")



if __name__ == "__main__":
    enviar_mensagens("agendamentos (5) (2).xlsx")
