import openpyxl
import re
import requests
from time import sleep

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# BUSCA ENDEREÇO PELO CEP 

def buscar_endereco_por_cep(cep):
    cep = re.sub(r"\D", "", str(cep))

    if len(cep) != 8:
        return None

    try:
        r = requests.get(f"https://viacep.com.br/ws/{cep}/json/", timeout=10)
        dados = r.json()

        cidade = dados.get("localidade", "")
        uf = dados.get("uf", "")
        bairro = dados.get("bairro")
        if not bairro or not bairro.strip():
            bairro = cidade

        return {
            "bairro": bairro,
            "cidade": cidade,
            "uf": uf
        }

    except:
        return None

options = Options()
options.add_argument("--start-maximized")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

# Abre o WhatsApp Web
driver.get("https://web.whatsapp.com/")
sleep(30)  # escanear QR Code uma vez

def enviar_mensagem(numero, mensagem):
    driver.get(f"https://web.whatsapp.com/send?phone={numero}")

    caixa = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located(
            (By.XPATH, "//footer//div[@role='textbox']")
        )
    )

    caixa.click()
    sleep(1)

    linhas = mensagem.split("\n")

    for i, linha in enumerate(linhas):
        caixa.send_keys(linha)
        sleep(0.5)  # delay maior = estabilidade e não trava

        if i < len(linhas) - 1:
            caixa.send_keys(Keys.SHIFT, Keys.ENTER)
            sleep(0.3)

    caixa.send_keys(Keys.ENTER)
    sleep(4)



# LÊ A PLANILHA E ENVIA AS MENSAGENS
def enviar_mensagens(caminho_planilha):
    wb = openpyxl.load_workbook(caminho_planilha)
    ws = wb.active

    # Mapeia cabeçalhos da linha 1
    cab = {cell.value.strip().lower(): i for i, cell in enumerate(ws[1])}

    for linha in ws.iter_rows(min_row=2):
        nome = linha[cab["cliente"]].value
        telefone = linha[cab["telefone"]].value
        servico = linha[cab["serviço"]].value
        produto = linha[cab["produto"]].value
        logradouro = linha[cab["logradouro"]].value
        numero = linha[cab["número"]].value
        cep = linha[cab["cep"]].value

        if not nome or not telefone or not cep:
            continue

        telefone = re.sub(r"\D", "", str(telefone))
        cep = re.sub(r"\D", "", str(cep))

        endereco = buscar_endereco_por_cep(cep)

        if not endereco:
            print(f"❌ Endereço não encontrado para {nome}")
            continue

        bairro = endereco["bairro"]
        cidade = endereco["cidade"]
        uf = endereco["uf"]

        mensagem = (
            f"Olá {nome}, tudo bem?\n\n"
            f"Estamos entrando em contato referente ao serviço de {servico}.\n\n"
            f"Produto: {produto}\n\n"
            f"Endereço:\n"
            f"{logradouro}, nº {numero}\n"
            f"{bairro} - {cidade}/{uf}\n"
            f"CEP: {cep}\n\n"
            f"Poderia confirmar, por gentileza, se as informações estão corretas?\n\n"
            f"Atenciosamente."
        )

        try:
            enviar_mensagem(telefone, mensagem)
            print(f"✅ Mensagem enviada para {nome}")
            sleep(20)  # anti-ban
        except Exception as e:
            print(f"❌ Erro ao enviar para {nome}: {e}")

if __name__ == "__main__":
    enviar_mensagens("agendamentos (5) (2).xlsx")
